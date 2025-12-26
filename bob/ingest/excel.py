"""Excel document parser."""

from __future__ import annotations

from pathlib import Path

from bob.ingest.base import DocumentSection, ParsedDocument, Parser


class ExcelParser(Parser):
    """Parser for Excel documents (.xlsx, .xls)."""

    extensions = [".xlsx", ".xls"]

    def can_parse(self, path: Path) -> bool:
        """Check if this parser can handle the given file."""
        return path.suffix.lower() in self.extensions

    def parse(self, path: Path) -> ParsedDocument:
        """Parse an Excel document into sheet-based sections."""
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        sections: list[DocumentSection] = []
        full_content: list[str] = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows: list[str] = []

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                # Convert row to string representation
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(v.strip() for v in row_values):  # Skip empty rows
                    row_text = " | ".join(row_values)
                    rows.append(f"Row {row_idx}: {row_text}")

            if rows:
                sheet_content = "\n".join(rows)
                full_content.append(f"=== Sheet: {sheet_name} ===\n{sheet_content}")

                sections.append(
                    DocumentSection(
                        content=sheet_content,
                        locator_type="sheet",
                        locator_value={
                            "sheet_name": sheet_name,
                            "row_count": len(rows),
                        },
                    )
                )

        wb.close()

        content = "\n\n".join(full_content)
        return ParsedDocument(
            source_path=str(path),
            source_type="excel",
            content=content,
            sections=sections,
            title=path.stem,
            source_date=self.get_source_date(path, content),
        )
